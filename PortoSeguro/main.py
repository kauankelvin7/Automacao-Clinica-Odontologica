import csv
import glob
import os
import random
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import (InvalidSessionIdException,
                                        NoSuchElementException,
                                        StaleElementReferenceException,
                                        TimeoutException)
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# ======================== CONFIGURAÇÃO DE ENCODING ========================
# Configura o encoding UTF-8 para o console do Windows
if os.name == 'nt':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        # Python < 3.7
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# ======================== CONFIGURAÇÃO DE CORES PARA O CONSOLE ========================
class Cores:
    """Códigos de escape ANSI para formatação de texto no terminal."""
    RESET = '\033[0m'
    BOLD = '\033[1m'
    VERDE = '\033[92m'
    AMARELO = '\033[93m'
    VERMELHO = '\033[91m'
    AZUL = '\033[94m'
    CYAN = '\033[96m'
    MAGENTA = '\033[95m'
    BG_VERDE = '\033[102m\033[30m'

if os.name == 'nt':
    os.system('')

# ======================== FUNÇÕES DE INTERFACE DE USUÁRIO ========================
def exibir_banner():
    """Limpa a tela e exibe o banner principal do sistema."""
    os.system('cls' if os.name == 'nt' else 'clear')
    print(Cores.CYAN + Cores.BOLD)
    print("╔" + "═" * 78 + "╗")
    print("║" + " 🦷 SISTEMA DE AUTOMAÇÃO PORTO SEGURO 🦷 ".center(84) + "║")
    print("║" + f" Versão 2.0 | {datetime.now().year} ".center(84) + "║")
    print("╚" + "═" * 78 + "╝")
    print(Cores.RESET)

def exibir_mensagem_sucesso(mensagem):
    print(f"{Cores.VERDE}✓ {mensagem}{Cores.RESET}")

def exibir_mensagem_erro(mensagem):
    print(f"{Cores.VERMELHO}✗ {mensagem}{Cores.RESET}")

def exibir_mensagem_info(mensagem):
    print(f"{Cores.CYAN}ℹ {mensagem}{Cores.RESET}")

def exibir_mensagem_alerta(mensagem):
    print(f"{Cores.AMARELO}⚠ {mensagem}{Cores.RESET}")

# ======================== DADOS E MAPEAMENTOS ========================
MAPEAMENTO_PROCEDIMENTOS = {
    "1": {"codigo_completo": "82000468", "descricao": "CONTROLE DE HEMORRAGIA COM APLICACAO", "requer_dente": False},
    "2": {"codigo_completo": "82000484", "descricao": "CONTROLE DE HEMORRAGIA SEM APLICACAO", "requer_dente": False},
    "3": {"codigo_completo": "82000859", "descricao": "EXODONTIA DE RAIZ RESIDUAL", "requer_dente": True},
    "4": {"codigo_completo": "82000875", "descricao": "EXODONTIA SIMPLES DE PERMANENTE", "requer_dente": True},
    "5": {"codigo_completo": "82001022", "descricao": "INCISAO E DRENAGEM EXTRA-ORAL", "requer_dente": False},
    "6": {"codigo_completo": "82001030", "descricao": "INCISAO E DRENAGEM INTRA-ORAL", "requer_dente": False},
    "7": {"codigo_completo": "82001197", "descricao": "REDUCAO SIMPLES DE LUXACAO", "requer_dente": False},
    "8": {"codigo_completo": "82001251", "descricao": "REIMPLANTE DENTARIO COM CONTENCAO", "requer_dente": True},
    "9": {"codigo_completo": "82001308", "descricao": "REMOCAO DE DRENO EXTRA-ORAL", "requer_dente": False},
    "10": {"codigo_completo": "82001316", "descricao": "REMOCAO DE DRENO INTRA-ORAL", "requer_dente": False},
    "11": {"codigo_completo": "82001499", "descricao": "SUTURA DE FERIDA EM REGIAO BUCO-MAXILO-FACIAL", "requer_dente": False},
    "12": {"codigo_completo": "82001650", "descricao": "TRATAMENTO DE ALVEOLITE", "requer_dente": True},
    "13": {"codigo_completo": "83000089", "descricao": "EXODONTIA SIMPLES DE DECIDUO", "requer_dente": True},
    "14": {"codigo_completo": "85100048", "descricao": "COLAGEM DE FRAGMENTOS DENTARIOS", "requer_dente": True},
    "15": {"codigo_completo": "85100196", "descricao": "RESTAURACAO EM RESINA FOTOPOLIMERIZAVEL 1 FACE", "requer_dente": True},
    "16": {"codigo_completo": "85100200", "descricao": "RESTAURACAO EM RESINA FOTOPOLIMERIZAVEL 2 FACES", "requer_dente": True},
    "17": {"codigo_completo": "85100218", "descricao": "RESTAURACAO EM RESINA FOTOPOLIMERIZAVEL 3 FACES", "requer_dente": True},
    "18": {"codigo_completo": "85100226", "descricao": "RESTAURACAO EM RESINA FOTOPOLIMERIZAVEL 4 FACES", "requer_dente": True},
    "19": {"codigo_completo": "85200034", "descricao": "PULPECTOMIA", "requer_dente": True},
    "20": {"codigo_completo": "85200085", "descricao": "RESTAURACAO TEMPORARIA / TRATAMENTO EXPECTANTE", "requer_dente": True},
    "21": {"codigo_completo": "85300020", "descricao": "IMOBILIZACAO DENTARIA EM DENTES PERMANENTES", "requer_dente": False},
    "22": {"codigo_completo": "85300063", "descricao": "TRATAMENTO DE ABSCESSO PERIODONTAL AGUDO", "requer_dente": False},
    "23": {"codigo_completo": "85300080", "descricao": "TRATAMENTO DE PERICORONARITE", "requer_dente": True},
    "24": {"codigo_completo": "85400084", "descricao": "COROA PROVISORIA SEM PINO", "requer_dente": True},
    "25": {"codigo_completo": "85400467", "descricao": "RECIMENTACAO DE TRABALHOS PROTETICOS", "requer_dente": True}
}

# ======================== CLASSES DE GERENCIAMENTO ========================
class LogManager:
    """Gerencia a criação de logs de erros e relatórios de sucesso."""
    def __init__(self, pasta_base):
        self.pasta_logs = Path(pasta_base) / "logs"
        self.pasta_logs.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        self.arquivo_log_falhas = self.pasta_logs / f"log_falhas_{timestamp}.txt"
        self.falhas = []
        self._inicializar_log_falhas()
        
        self.arquivo_relatorio_sucesso = self.pasta_logs / f"relatorio_sucesso_{timestamp}.csv"
        self._inicializar_relatorio_sucesso()

    def _inicializar_log_falhas(self):
        with open(self.arquivo_log_falhas, 'w', encoding='utf-8') as f:
            f.write(f"LOG DE FALHAS - AUTOMAÇÃO PORTO SEGURO\nData/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'='*80}\n\n")

    def _inicializar_relatorio_sucesso(self):
        with open(self.arquivo_relatorio_sucesso, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Guia', 'Codigo_Planilha', 'Dente_Selecionado', 'Status', 'Data_Hora'])
            
    def registrar_sucesso(self, id_guia, cod_simplificado, dente_val):
        with open(self.arquivo_relatorio_sucesso, 'a', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([id_guia, cod_simplificado, dente_val, 'Sucesso', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])

    def registrar_falha(self, id_guia, motivo, traceback_erro=""):
        registro = {'id_guia': id_guia, 'motivo': motivo}
        if not any(f['id_guia'] == id_guia and f['motivo'] == motivo for f in self.falhas):
            self.falhas.append(registro)
            with open(self.arquivo_log_falhas, 'a', encoding='utf-8') as f:
                f.write(f"\n{'─' * 80}\nGUIA: {id_guia}\nData/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\nMotivo: {motivo}\n")
                if traceback_erro: f.write(f"\nTraceback:\n{traceback_erro}\n")
                f.write(f"{'─' * 80}\n")

    def gerar_resumos(self):
        exibir_mensagem_info(f"Relatório de sucessos salvo em: {self.arquivo_relatorio_sucesso}")
        if self.falhas:
            exibir_mensagem_info(f"Log de falhas salvo em: {self.arquivo_log_falhas}")
            exibir_mensagem_alerta(f"{len(self.falhas)} guia(s) necessitam de verificação manual.")

class PortoSeguroBot:
    """Encapsula toda a lógica de interação com o portal Porto Seguro Odonto."""
    def __init__(self, tipo_usuario="completo"):
        self.url = "https://www.dentistaportoseguro.com.br/"
        
        # Configuração de usuários
        self.usuarios = {
            "completo": {
                "login": "00281027",
                "senha": "@Isbi4420",
                "nome": "Usuário Completo (com Procedimento Complementar)",
                "usa_modal_complementar": True
            },
            "simplificado": {
                "login": "00260115",
                "senha": "@Isbi4420",
                "nome": "Usuário Simplificado (sem Procedimento Complementar)",
                "usa_modal_complementar": False
            }
        }
        
        # Define o tipo de usuário atual
        self.tipo_usuario = tipo_usuario
        self.login = self.usuarios[tipo_usuario]["login"]
        self.senha = self.usuarios[tipo_usuario]["senha"]
        self.usa_modal_complementar = self.usuarios[tipo_usuario]["usa_modal_complementar"]
        
        self.driver = None
        self.wait = None
        self.log_manager = None

    def _human_delay(self, min_seconds=0.5, max_seconds=1.0):
        """Pausa a execução por um tempo aleatório para simular comportamento humano."""
        time.sleep(random.uniform(min_seconds, max_seconds))
    
    def _pause_between_actions(self, seconds=6):
        """Pausa estratégica entre ações importantes do fluxo."""
        exibir_mensagem_info(f"Aguardando {seconds} segundos antes da próxima ação...")
        time.sleep(seconds)

    def handle_overlays(self, main_timeout=30):
        """Aguarda o desaparecimento de pop-ups de carregamento (spinners)."""
        try:
            WebDriverWait(self.driver, 2).until(EC.visibility_of_element_located((By.ID, "spin_modal_overlay")))
            exibir_mensagem_info("Overlay detectado. Aguardando...")
            WebDriverWait(self.driver, main_timeout).until(EC.invisibility_of_element_located((By.ID, "spin_modal_overlay")))
        except TimeoutException:
            pass

    def iniciar_driver(self):
        """Configura e inicia o WebDriver do Selenium."""
        exibir_mensagem_info("Iniciando navegador Chrome...")
        options = webdriver.ChromeOptions()
        options.add_argument('--start-maximized')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        self.wait = WebDriverWait(self.driver, 20)
        exibir_mensagem_sucesso("Driver iniciado com sucesso")

    def fazer_login(self):
        """Executa o processo de login no portal."""
        exibir_mensagem_info("Acessando portal Porto Seguro...")
        self.driver.get(self.url)
        self._human_delay(1, 2)
        self.driver.find_element(By.ID, "usuario").send_keys(self.login)
        self._human_delay(0.5, 1)
        self.driver.find_element(By.ID, "senha").send_keys(self.senha)
        self._human_delay(0.5, 1)
        self.driver.find_element(By.ID, "login-submit").click()
        self.handle_overlays()
        self._pause_between_actions(2)
        exibir_mensagem_sucesso("Login realizado com sucesso")

    def navegar_meus_tratamentos(self):
        """Navega para a página 'Meus Tratamentos'."""
        try:
            self._pause_between_actions(2)
            self.handle_overlays()
            menu = self.wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Meus Tratamentos")))
            self.driver.execute_script("arguments[0].click();", menu)
            self.handle_overlays()
            self._human_delay(2, 3)
        except Exception as e:
            exibir_mensagem_erro(f"Erro ao navegar para 'Meus Tratamentos': {e}")
            raise

    def buscar_guia(self, id_guia):
        """Busca uma guia específica na página 'Meus Tratamentos'."""
        try:
            self.handle_overlays()
            campo_busca = self.wait.until(EC.presence_of_element_located((By.ID, "searchGTO")))
            # Otimização: insere o valor via JS para maior velocidade
            self.driver.execute_script("arguments[0].value = arguments[1];", campo_busca, id_guia)
            self._human_delay(0.5, 1)
            campo_busca.send_keys(Keys.ENTER)
            self.handle_overlays()
            self._human_delay(1, 2)
            link_guia = self.wait.until(EC.element_to_be_clickable((By.XPATH, f"//a[text()='{id_guia}']")))
            self.driver.execute_script("arguments[0].click();", link_guia)
            self.handle_overlays()
            self._pause_between_actions(5)
            exibir_mensagem_sucesso(f"Guia {id_guia} aberta")
            return True
        except TimeoutException:
            exibir_mensagem_erro(f"Guia {id_guia} não encontrada")
            return False

    def verificar_procedimento_ja_confirmado(self):
        """Verifica se o procedimento já foi confirmado (ícone de dólar laranja #de9045) - significa que precisa apenas anexar arquivos."""
        try:
            self.handle_overlays()
            # Procura ESPECIFICAMENTE por ícones fa-dollar que tenham a cor no style
            icones_dollar = self.driver.find_elements(By.XPATH, "//i[contains(@class, 'fa-dollar') and contains(@style, 'color')]")
            
            exibir_mensagem_info(f"Debug - Total de ícones $ com estilo encontrados: {len(icones_dollar)}")
            
            # Se não encontrou nenhum ícone, retorna False
            if not icones_dollar:
                exibir_mensagem_info("✓ Nenhum ícone $ com estilo encontrado - guia nova, processamento normal.")
                return False
            
            # Verifica TODOS os ícones encontrados para achar um laranja
            for idx, icone_dollar in enumerate(icones_dollar):
                # Verifica se a cor do ícone é #de9045 (laranja/dourado) usando JavaScript
                cor_icone = self.driver.execute_script(
                    "return window.getComputedStyle(arguments[0]).getPropertyValue('color');", 
                    icone_dollar
                )
                
                # Também verifica o atributo style direto
                style_attr = icone_dollar.get_attribute('style') or ""
                
                # Debug: mostra a cor de cada ícone encontrado
                exibir_mensagem_info(f"Debug - Ícone ${idx+1}: Cor={cor_icone} | Style={style_attr[:80]}...")
                
                # Verifica se tem a cor #de9045 no style ou se a cor computada é rgb(222, 144, 69) que é equivalente a #de9045
                tem_cor_laranja = '#de9045' in style_attr.lower() or 'rgb(222, 144, 69)' in cor_icone
                
                if tem_cor_laranja:
                    exibir_mensagem_info("✓ Ícone $ LARANJA detectado - procedimento já confirmado. Indo direto para anexos.")
                    return True
            
            # Se chegou aqui, encontrou ícones $ mas nenhum é laranja
            exibir_mensagem_info("✗ Ícones $ encontrados mas NENHUM é laranja - guia precisa ser processada normalmente.")
            return False
        except Exception as e:
            exibir_mensagem_info(f"Sem ícone $ detectado - guia nova ou não iniciada. (Erro: {str(e)[:50]})")
            return False

    def verificar_guia_ja_finalizada(self):
        """Verifica se a guia já foi completamente finalizada e paga (ícone de dólar verde ou outra cor que indique conclusão total)."""
        try:
            self.handle_overlays()
            # Procura ESPECIFICAMENTE por ícones fa-dollar que tenham a cor no style
            icones_dollar = self.driver.find_elements(By.XPATH, "//i[contains(@class, 'fa-dollar') and contains(@style, 'color')]")
            
            # Se não encontrou nenhum ícone, retorna False (guia não finalizada)
            if not icones_dollar:
                return False
            
            # Verifica o primeiro ícone encontrado
            icone_dollar = icones_dollar[0]
            
            # Verifica a cor do ícone
            cor_icone = self.driver.execute_script(
                "return window.getComputedStyle(arguments[0]).getPropertyValue('color');", 
                icone_dollar
            )
            style_attr = icone_dollar.get_attribute('style') or ""
            
            # Se NÃO for laranja (#de9045), considera como finalizada
            nao_e_laranja = '#de9045' not in style_attr.lower() and 'rgb(222, 144, 69)' not in cor_icone
            
            if nao_e_laranja:
                exibir_mensagem_info(f"✓ Guia já completamente finalizada (ícone $ detectado, cor: {cor_icone}).")
                return True
            return False
        except Exception:
            return False

    def verificar_ja_esta_na_tela_anexos(self):
        """Verifica se já está na tela de anexos."""
        try:
            self.handle_overlays()
            # Verifica se o título "Gestão Rede" está presente ou o botão de anexos
            titulo_presente = len(self.driver.find_elements(By.XPATH, "//title[contains(text(), 'Gestão Rede')]")) > 0
            botao_anexos_presente = len(self.driver.find_elements(By.ID, "btnAnexosProcedimento")) > 0
            
            if titulo_presente or botao_anexos_presente:
                exibir_mensagem_info("Detectado que a guia já está na tela de anexos. Pulando confirmação de procedimento.")
                return True
            return False
        except Exception:
            return False

    def _extrair_data_autorizacao(self):
        """Extrai a data de autorização da página da guia."""
        try:
            self.handle_overlays()
            elemento_p = self.wait.until(EC.presence_of_element_located((By.XPATH, "//p[strong[text()='Data da Autorização: ']]")))
            match = re.search(r'\d{2}/\d{2}/\d{4}', elemento_p.text)
            return match.group(0) if match else None
        except TimeoutException:
            exibir_mensagem_alerta("Data de autorização não encontrada")
            return None

    def _set_and_verify_date(self, date_element, date_to_set, checkbox):
        """Preenche e verifica um campo de data."""
        try:
            self.handle_overlays()
            self.driver.execute_script("arguments[0].checked=true; arguments[1].removeAttribute('disabled'); $(arguments[0]).trigger('change');", checkbox, date_element)
            date_element.send_keys(Keys.CONTROL, "a", Keys.BACKSPACE)
            date_element.send_keys(date_to_set)
            ActionChains(self.driver).move_to_element(self.driver.find_element(By.ID, "observacaoJustificativa")).click().perform() # Tira o foco
            self.handle_overlays()
            return date_element.get_attribute('value') == date_to_set
        except Exception as e:
            exibir_mensagem_erro(f"Erro ao inserir data: {e}")
            return False
        
    def _handle_faturamento_guia_modal(self, timeout=3):
        """Lida com o modal 'Faturamento através de' se ele aparecer."""
        try:
            botao_guia = WebDriverWait(self.driver, timeout).until(EC.element_to_be_clickable((By.ID, "btnGuia")))
            self.driver.execute_script("arguments[0].click();", botao_guia)
            self.handle_overlays(main_timeout=15)
        except TimeoutException:
            pass

    def _verificar_se_requer_dente(self, codigo_procedimento):
        """Verifica se um procedimento requer seleção de dente baseado no mapeamento."""
        for key, dados in MAPEAMENTO_PROCEDIMENTOS.items():
            if dados["codigo_completo"] == codigo_procedimento:
                requer = dados.get("requer_dente", True)  # Default: True (requer dente)
                return requer
        # Se não encontrar no mapeamento, assume que requer dente por segurança
        exibir_mensagem_alerta(f"Procedimento {codigo_procedimento} não encontrado no mapeamento. Assumindo que requer dente.")
        return True        

    def _handle_procedimento_complementar(self, proc_code_alvo, proc_desc_alvo, dente_valor_alvo, observacao_alvo="", codigo_guia=""):
        """Lida com o modal 'Procedimento Complementar' com validação rigorosa."""
        try:
            exibir_mensagem_info("Aguardando modal de procedimento complementar...")
            self.handle_overlays()
            wait = WebDriverWait(self.driver, 15)
            
            # Verifica se o procedimento requer seleção de dente
            requer_dente = self._verificar_se_requer_dente(proc_code_alvo)
            
            # Espera diretamente pela célula-alvo para garantir que todo o conteúdo do modal carregou.
            celula_procedimento_xpath = f"//td[starts-with(normalize-space(.), '{proc_code_alvo} -')]"
            celula_procedimento = wait.until(EC.presence_of_element_located((By.XPATH, celula_procedimento_xpath)))
            exibir_mensagem_sucesso(f"Modal e procedimento '{proc_code_alvo}' encontrados. Validando...")

            linha = celula_procedimento.find_element(By.XPATH, "./..")
            texto_procedimento_site = linha.find_element(By.XPATH, ".//td[2]").text.upper()
            
            if proc_desc_alvo.upper() not in texto_procedimento_site:
                exibir_mensagem_erro(f"VALIDAÇÃO FALHOU! O nome do procedimento '{proc_code_alvo}' não corresponde.")
                return False
            exibir_mensagem_sucesso("Validação de nome OK.")
            
            checkbox_proc = linha.find_element(By.XPATH, ".//input[@type='checkbox']")
            self.driver.execute_script("arguments[0].click();", checkbox_proc)
            self.handle_overlays()
            self._human_delay(0.5, 1)

            # Só seleciona o dente se o procedimento requerer
            if requer_dente and dente_valor_alvo:
                exibir_mensagem_info(f"Selecionando dente {dente_valor_alvo} para procedimento {proc_code_alvo}...")
                Select(linha.find_element(By.XPATH, ".//select")).select_by_value(dente_valor_alvo)
                self.handle_overlays()
                self._human_delay(0.5, 1)
            else:
                exibir_mensagem_info(f"Procedimento {proc_code_alvo} NÃO requer seleção de dente. Pulando...")

            if observacao_alvo and codigo_guia:
                try:
                    codigo_base = codigo_guia[:8] if len(codigo_guia) >= 8 else codigo_guia
                    campo_obs_id = f"obs_{proc_code_alvo}_{codigo_base}"
                    exibir_mensagem_info(f"Tentando preencher observação no campo '{campo_obs_id}'...")
                    
                    campo_obs = self.driver.find_element(By.ID, campo_obs_id)
                    campo_obs.clear()
                    campo_obs.send_keys(observacao_alvo)
                    self._human_delay(0.5, 1)
                    exibir_mensagem_sucesso(f"Observação inserida: '{observacao_alvo[:50]}...'")
                except NoSuchElementException:
                    exibir_mensagem_alerta(f"Campo de observação '{campo_obs_id}' não encontrado. Continuando sem observação.")
                except Exception as e:
                    exibir_mensagem_alerta(f"Erro ao preencher observação: {e}. Continuando...")

            self.driver.find_element(By.ID, "btnAvancarAlteracao").click()
            self.handle_overlays()
            time.sleep(9)
            exibir_mensagem_sucesso("Procedimento complementar salvo.")
            return True
        except (TimeoutException, NoSuchElementException):
            exibir_mensagem_erro(f"Falha Crítica: O procedimento com código '{proc_code_alvo}' não foi encontrado no modal dentro de 15s.")
            exibir_mensagem_alerta("Causa provável: O código na planilha está incorreto ou não existe neste modal.")
            return False
        except Exception as e:
            exibir_mensagem_erro(f"Erro inesperado ao processar o modal complementar: {e}")
            self.log_manager.registrar_falha(proc_code_alvo, "Erro no modal complementar", traceback.format_exc())
            return False


    def confirmar_realizacao_procedimento(self, proc_code, proc_desc, dente_valor, observacao="", codigo_guia=""):
        """Confirma a data de realização e lida com os modais subsequentes."""
        exibir_mensagem_info("Confirmando realização do procedimento...")
        self._pause_between_actions(4)
        data_autorizacao = self._extrair_data_autorizacao()
        if not data_autorizacao: return False
        try:
            checkboxes = self.driver.find_elements(By.XPATH, "//input[contains(@id, 'chk_atualiza_data_')]")
            if checkboxes:
                for chk in checkboxes:
                    proc_id_chk = chk.get_attribute('id').split('_')[-1]
                    campo_data = self.driver.find_element(By.ID, f"dt_conf_{proc_id_chk}")
                    if not self._set_and_verify_date(campo_data, data_autorizacao, chk): return False
            
            self._human_delay(1, 2)
            self.wait.until(EC.element_to_be_clickable((By.ID, "btnConfirmar"))).click()
            self.handle_overlays()
            self._pause_between_actions(6)
            
            # Verifica se deve processar o modal complementar baseado no tipo de usuário
            if self.usa_modal_complementar:
                if proc_code and dente_valor:
                    if not self._handle_procedimento_complementar(proc_code, proc_desc, dente_valor, observacao, codigo_guia):
                        return False
            else:
                exibir_mensagem_info("Usuário simplificado: pulando modal de Procedimento Complementar.")
            
            self._pause_between_actions(10)
            self._handle_faturamento_guia_modal()
            self._pause_between_actions(6)
            exibir_mensagem_sucesso("Etapa de confirmação concluída.")
            return True
        except Exception as e:
            exibir_mensagem_erro(f"Erro ao confirmar procedimento: {e}")
            return False

    def buscar_arquivos_guia(self, id_guia, pasta_base):
        """Busca arquivos locais correspondentes a uma guia."""
        exibir_mensagem_info(f"Buscando arquivos para guia {id_guia}...")
        arquivos_encontrados = []
        id_parcial = id_guia[-4:]
        for f in Path(pasta_base).rglob('*'):
            if not f.is_file(): continue
            nome_upper = f.stem.upper()
            if id_guia in nome_upper or id_parcial in nome_upper:
                tipo_arquivo = 'documentacao'
                if nome_upper.startswith('FI'): tipo_arquivo = 'foto-inicial'
                elif nome_upper.startswith('FF'): tipo_arquivo = 'foto-final'
                elif 'RX' in nome_upper or nome_upper.startswith(('RF', 'RI')): tipo_arquivo = 'raio-x'
                elif 'LAUDO' in nome_upper: tipo_arquivo = 'laudo'
                elif nome_upper.startswith(('G', 'GUIA', 'ASS')): tipo_arquivo = 'guia'
                arquivos_encontrados.append({'caminho': str(f.resolve()), 'tipo': tipo_arquivo})
                print(f"  {Cores.VERDE}→{Cores.RESET} {f.name} (tipo: {tipo_arquivo})")
        return arquivos_encontrados

    def _abrir_modal_anexos(self):
        """Aguarda a página estabilizar e abre o modal de anexos."""
        exibir_mensagem_info("Preparando para anexar arquivos...")
        try:
            time.sleep(5)
            self.handle_overlays()
            botao_anexos = self.wait.until(EC.element_to_be_clickable((By.ID, "btnAnexosProcedimento")))
            time.sleep(2) 
            self.driver.execute_script("arguments[0].click();", botao_anexos)
            time.sleep(5)
            self.handle_overlays()
            self.wait.until(EC.visibility_of_element_located((By.ID, "modalAnexosProcedimento")))
            exibir_mensagem_sucesso("Modal de anexos aberto.")
        except Exception as e:
            exibir_mensagem_erro(f"Não foi possível abrir o modal de anexos de forma estável: {e}")
            raise

    def _fechar_modal_anexos(self):
        """Fecha o modal de anexos."""
        try:
            self.handle_overlays()
            self.driver.find_element(By.XPATH, "//div[@id='modalAnexosProcedimento' and contains(@style,'display: block')]//button[@data-dismiss='modal']").click()
            self.wait.until(EC.invisibility_of_element_located((By.ID, "modalAnexosProcedimento")))
            self.handle_overlays()
        except Exception: pass

    def _anexar_um_arquivo(self, anexo_info):
        """Executa a lógica de upload para um único arquivo."""
        caminho_arquivo, tipo = anexo_info['caminho'], anexo_info['tipo']
        print(f"  {Cores.CYAN}→{Cores.RESET} Anexando: {Path(caminho_arquivo).name}")
        
        self.handle_overlays()
        select_tipo = Select(self.wait.until(EC.presence_of_element_located((By.ID, "tipoAnexoModal"))))
        if tipo == 'raio-x': 
            select_tipo.select_by_value("1")
        elif tipo == 'laudo': 
            select_tipo.select_by_value("2")
        elif tipo == 'guia': 
            select_tipo.select_by_value("3")
        elif tipo in ['foto-inicial', 'foto-final', 'documentacao']:
            select_tipo.select_by_value("4") 
        else: 
            select_tipo.select_by_value("4")  # Default: Documentação (Fotos)
        
        self.handle_overlays()
        self._human_delay(0.3, 0.5)

        if tipo == 'raio-x':
            try: 
                WebDriverWait(self.driver, 3).until(EC.element_to_be_clickable((By.XPATH, "//label[@for='periapical']"))).click()
                self.handle_overlays()
            except TimeoutException: pass
        
        Select(self.driver.find_element(By.ID, "processoAnexo")).select_by_value("2")
        self.handle_overlays()
        self._human_delay(0.3, 0.5)
        
        self.driver.find_element(By.ID, "arquivoAnexo").send_keys(caminho_arquivo)
        self._human_delay(0.5, 1)
        self.handle_overlays()
        
        for checkbox in self.driver.find_elements(By.XPATH, "//table[@id='tblVincularProcedimentos']//input[@type='checkbox']"):
            if not checkbox.is_selected(): checkbox.click()
        
        self.handle_overlays()
        self.driver.find_element(By.ID, "btnIncluirAnexos").click()
        self.handle_overlays(main_timeout=15)

    def anexar_todos_os_arquivos(self, lista_de_anexos):
        """Itera sobre uma lista de arquivos e anexa cada um."""
        if not lista_de_anexos:
            exibir_mensagem_alerta("Nenhum arquivo encontrado para anexar.")
            return True
        exibir_mensagem_info(f"Anexando {len(lista_de_anexos)} arquivo(s)...")
        self._pause_between_actions(4)
        try:
            for anexo in lista_de_anexos:
                self._abrir_modal_anexos()
                self._anexar_um_arquivo(anexo)
                self._fechar_modal_anexos()
                self._human_delay(1, 2)
            exibir_mensagem_sucesso("Todos os arquivos foram anexados.")
            self._pause_between_actions(8)
            return True
        except Exception as e:
            exibir_mensagem_erro(f"Erro durante o anexo de arquivos: {e}")
            self.log_manager.registrar_falha("N/A", "Erro no anexo", traceback.format_exc())
            return False

    def submeter_para_pagamento(self):
        """Clica no botão final para submeter a guia para pagamento."""
        exibir_mensagem_info("Submetendo guia para pagamento...")
        self._pause_between_actions(4)
        try:
            # Marca todos os checkboxes que aparecerem antes da confirmação final
            self.handle_overlays()
            checkboxes = self.driver.find_elements(By.XPATH, "//input[@type='checkbox']")
            
            if checkboxes:
                exibir_mensagem_info(f"Marcando {len(checkboxes)} checkbox(es) antes da confirmação...")
                for checkbox in checkboxes:
                    try:
                        if not checkbox.is_selected():
                            self.driver.execute_script("arguments[0].click();", checkbox)
                            self._human_delay(0.3, 0.5)
                    except Exception as e:
                        pass
                exibir_mensagem_sucesso(f"Checkbox(es) marcado(s).")
            
            self.handle_overlays()
            self._human_delay(1, 2)
            
            # Clica no botão de confirmar
            self.wait.until(EC.element_to_be_clickable((By.ID, "btnConfirmar"))).click()
            self.handle_overlays()
            self._pause_between_actions(6)
            self.wait.until(EC.presence_of_element_located((By.XPATH, "//i[contains(@class, 'fa-dollar')]")))
            exibir_mensagem_sucesso("Guia submetida e pagamento confirmado ($)")
            return True
        except Exception as e:
            exibir_mensagem_erro(f"Erro na submissão final: {e}")
            return False

# ======================== FUNÇÃO DE ORQUESTRAÇÃO PRINCIPAL ========================
def processar_guias_da_planilha(bot, caminho_planilha, pasta_arquivos_busca):
    """Orquestra todo o processo de automação, desde a leitura da planilha até o final."""
    try:
        workbook = openpyxl.load_workbook(caminho_planilha)
        sheet = workbook.active
    except FileNotFoundError:
        exibir_mensagem_erro(f"Planilha não encontrada: {caminho_planilha}")
        return

    guias_a_processar = []
    print(f"\n{Cores.CYAN}Lendo planilha... Total de linhas: {sheet.max_row}{Cores.RESET}")
    
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_col=4, values_only=True), 1):
        if row[0]:
            # Converte o código e remove decimais desnecessários (ex: 6.0 -> 6)
            cod_simplificado = ""
            if row[1]:
                cod_raw = str(row[1]).strip()
                # Remove .0 se for número decimal
                if '.' in cod_raw and cod_raw.replace('.', '').replace('-', '').isdigit():
                    cod_simplificado = str(int(float(cod_raw)))
                else:
                    cod_simplificado = cod_raw
            
            if cod_simplificado and cod_simplificado not in MAPEAMENTO_PROCEDIMENTOS:
                exibir_mensagem_erro(f"ERRO na linha {i} da planilha: Código '{cod_simplificado}' não foi encontrado no dicionário.")
                continue

            # Converte guia removendo decimais se necessário
            guia_raw = str(row[0]).strip()
            if '.' in guia_raw and guia_raw.replace('.', '').replace('-', '').isdigit():
                guia_formatada = str(int(float(guia_raw)))
            else:
                guia_formatada = guia_raw
            
            # Converte dente removendo decimais se necessário
            dente_raw = str(row[2]).strip() if row[2] else ""
            if dente_raw and '.' in dente_raw and dente_raw.replace('.', '').replace('-', '').isdigit():
                dente_formatado = str(int(float(dente_raw)))
            else:
                dente_formatado = dente_raw

            guia_dict = {
                'guia': guia_formatada,
                'cod_simplificado': cod_simplificado,
                'dente': dente_formatado,
                'observacao': str(row[3]).strip() if row[3] else ""
            }
            guias_a_processar.append(guia_dict)
    
    print(f"{Cores.VERDE}Total de guias válidas encontradas: {len(guias_a_processar)}{Cores.RESET}")
    
    total_guias = len(guias_a_processar)
    exibir_banner()
    print(f"\n{Cores.BOLD}{Cores.AZUL}{'═' * 80}{Cores.RESET}")
    print(f"{Cores.BOLD}INICIANDO PROCESSAMENTO DE {total_guias} GUIAS{Cores.RESET}")
    
    sucessos = 0
    try:
        bot.iniciar_driver()
        bot.fazer_login()
    except Exception as e:
        exibir_mensagem_erro(f"Falha crítica na inicialização: {e}")
        bot.log_manager.registrar_falha("GERAL", "Falha no Login/Inicialização", traceback.format_exc())
        return

    for i, guia_info in enumerate(guias_a_processar, 1):
        id_guia = guia_info['guia']
        cod_simplificado = guia_info['cod_simplificado']
        dente_val = guia_info['dente']

        proc_info = MAPEAMENTO_PROCEDIMENTOS.get(cod_simplificado, {})
        proc_code = proc_info.get('codigo_completo', "")
        proc_desc = proc_info.get('descricao', "")
        
        print(f"\n{Cores.MAGENTA}{'─' * 80}")
        print(f"► Processando Guia {i}/{total_guias}: {Cores.BOLD}{id_guia}{Cores.RESET}")
        
        try:
            bot.navegar_meus_tratamentos()
            if not bot.buscar_guia(id_guia):
                bot.log_manager.registrar_falha(id_guia, "Guia não encontrada")
                continue
            
            # Verifica se a guia já está completamente finalizada (não precisa fazer nada)
            if bot.verificar_guia_ja_finalizada():
                exibir_mensagem_sucesso("Guia já completamente finalizada. Pulando.")
                sucessos += 1
                bot.log_manager.registrar_sucesso(id_guia, "N/A", "N/A - Já finalizada")
                continue

            # Verifica se o procedimento já foi confirmado (ícone $ laranja)
            procedimento_ja_confirmado = bot.verificar_procedimento_ja_confirmado()
            
            # Se o procedimento ainda não foi confirmado, confirma agora
            if not procedimento_ja_confirmado:
                observacao = guia_info.get('observacao', "")
                if not bot.confirmar_realizacao_procedimento(proc_code, proc_desc, dente_val, observacao, id_guia):
                    bot.log_manager.registrar_falha(id_guia, "Falha ao confirmar procedimento ou na validação do modal")
                    continue
            else:
                exibir_mensagem_sucesso("Procedimento já confirmado ($ laranja). Indo direto para anexos...")
                # Aguarda a página estabilizar antes de anexar
                bot._pause_between_actions(3)

            # Anexa arquivos
            lista_de_anexos = bot.buscar_arquivos_guia(id_guia, pasta_arquivos_busca)
            if not bot.anexar_todos_os_arquivos(lista_de_anexos):
                bot.log_manager.registrar_falha(id_guia, "Falha no anexo de arquivos")
                continue

            # Submete para pagamento final
            if not bot.submeter_para_pagamento():
                bot.log_manager.registrar_falha(id_guia, "Falha na submissão final")
                continue
            
            print(f"\n{Cores.BG_VERDE} ✓ GUIA {id_guia} FINALIZADA COM SUCESSO {Cores.RESET}\n")
            sucessos += 1
            
            bot.log_manager.registrar_sucesso(id_guia, cod_simplificado, dente_val)

        except KeyboardInterrupt:
            exibir_mensagem_alerta("\nProcessamento interrompido pelo usuário!")
            break
        except Exception as e:
            exibir_mensagem_erro(f"Erro inesperado na guia {id_guia}: {e}")
            bot.log_manager.registrar_falha(id_guia, "Erro inesperado", traceback.format_exc())

    print(f"\n{Cores.CYAN}{'═' * 80}\n{'PROCESSAMENTO FINALIZADO'.center(80)}\n{'═' * 80}{Cores.RESET}")
    print(f"{Cores.BOLD}Total de guias na planilha:{Cores.RESET} {total_guias}")
    print(f"{Cores.VERDE}Processadas com sucesso:{Cores.RESET} {sucessos}")
    print(f"{Cores.VERMELHO}Guias com falha:{Cores.RESET} {len(bot.log_manager.falhas)}")
    print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")

# ======================== FUNÇÕES DO MENU ========================
def listar_logs_disponiveis(pasta_logs):
    """Lista todos os arquivos de log disponíveis."""
    if not pasta_logs.exists():
        return [], []
    
    logs_falha = sorted(pasta_logs.glob("log_falhas_*.txt"), key=os.path.getmtime, reverse=True)
    logs_sucesso = sorted(pasta_logs.glob("relatorio_sucesso_*.csv"), key=os.path.getmtime, reverse=True)
    
    return logs_falha, logs_sucesso

def exibir_log(caminho_log):
    """Exibe o conteúdo de um arquivo de log."""
    try:
        with open(caminho_log, 'r', encoding='utf-8') as f:
            conteudo = f.read()
        
        exibir_banner()
        print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}")
        print(f"{Cores.BOLD}📄 {caminho_log.name}{Cores.RESET}")
        print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
        print(conteudo)
        print(f"\n{Cores.CYAN}{'═' * 80}{Cores.RESET}")
    except Exception as e:
        exibir_mensagem_erro(f"Erro ao ler o log: {e}")

def menu_visualizar_logs(pasta_base):
    """Menu para visualizar logs anteriores."""
    pasta_logs = pasta_base / "logs"
    
    while True:
        exibir_banner()
        print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}")
        print(f"{Cores.BOLD}{'📊 VISUALIZAR LOGS'.center(80)}{Cores.RESET}")
        print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
        
        logs_falha, logs_sucesso = listar_logs_disponiveis(pasta_logs)
        
        if not logs_falha and not logs_sucesso:
            exibir_mensagem_alerta("Nenhum log encontrado.")
            input("\nPressione ENTER para voltar ao menu principal...")
            return
        
        print(f"{Cores.AMARELO}📋 LOGS DE FALHAS:{Cores.RESET}")
        if logs_falha:
            for idx, log in enumerate(logs_falha[:10], 1):
                data_mod = datetime.fromtimestamp(os.path.getmtime(log)).strftime('%d/%m/%Y %H:%M:%S')
                print(f"  {idx}. {log.name} - {Cores.CYAN}{data_mod}{Cores.RESET}")
        else:
            print(f"  {Cores.VERDE}Nenhum log de falha encontrado{Cores.RESET}")
        
        print(f"\n{Cores.VERDE}✓ LOGS DE SUCESSO:{Cores.RESET}")
        offset = len(logs_falha[:10])
        if logs_sucesso:
            for idx, log in enumerate(logs_sucesso[:10], offset + 1):
                data_mod = datetime.fromtimestamp(os.path.getmtime(log)).strftime('%d/%m/%Y %H:%M:%S')
                print(f"  {idx}. {log.name} - {Cores.CYAN}{data_mod}{Cores.RESET}")
        else:
            print(f"  {Cores.AMARELO}Nenhum log de sucesso encontrado{Cores.RESET}")
        
        print(f"\n{Cores.CYAN}{'─' * 80}{Cores.RESET}")
        print(f"  0. {Cores.VERMELHO}← Voltar ao menu principal{Cores.RESET}")
        print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
        
        try:
            escolha = input(f"{Cores.BOLD}Digite o número do log que deseja visualizar: {Cores.RESET}").strip()
            
            if escolha == "0":
                return
            
            escolha_num = int(escolha)
            todos_logs = logs_falha[:10] + logs_sucesso[:10]
            
            if 1 <= escolha_num <= len(todos_logs):
                exibir_log(todos_logs[escolha_num - 1])
                input(f"\n{Cores.CYAN}Pressione ENTER para voltar...{Cores.RESET}")
            else:
                exibir_mensagem_erro("Número inválido!")
                time.sleep(1.5)
        except ValueError:
            exibir_mensagem_erro("Digite um número válido!")
            time.sleep(1.5)
        except KeyboardInterrupt:
            return

def exibir_menu_principal():
    """Exibe o menu principal do sistema."""
    exibir_banner()
    print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}")
    print(f"{Cores.BOLD}{'📋 MENU PRINCIPAL'.center(80)}{Cores.RESET}")
    print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
    print(f"  {Cores.VERDE}1.{Cores.RESET} {Cores.BOLD}🚀 Iniciar Processamento de Guias{Cores.RESET}")
    print(f"  {Cores.AZUL}2.{Cores.RESET} {Cores.BOLD}📊 Visualizar Logs Anteriores{Cores.RESET}")
    print(f"  {Cores.AMARELO}3.{Cores.RESET} {Cores.BOLD}ℹ️  Informações do Sistema{Cores.RESET}")
    print(f"  {Cores.VERMELHO}0.{Cores.RESET} {Cores.BOLD}❌ Sair{Cores.RESET}")
    print(f"\n{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")

def selecionar_tipo_usuario():
    """Menu para selecionar o tipo de usuário."""
    while True:
        exibir_banner()
        print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}")
        print(f"{Cores.BOLD}{'👤 SELEÇÃO DE USUÁRIO'.center(80)}{Cores.RESET}")
        print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
        
        print(f"  {Cores.VERDE}1.{Cores.RESET} {Cores.BOLD}Usuário Completo{Cores.RESET}")
        print(f"     {Cores.CYAN}└─ Login: 00281027{Cores.RESET}")
        print(f"     {Cores.CYAN}└─ Com Procedimento Complementar{Cores.RESET}")
        print(f"     {Cores.CYAN}└─ Validação de códigos e seleção de dentes{Cores.RESET}\n")
        
        print(f"  {Cores.AZUL}2.{Cores.RESET} {Cores.BOLD}Usuário Simplificado{Cores.RESET}")
        print(f"     {Cores.CYAN}└─ Login: 00260115{Cores.RESET}")
        print(f"     {Cores.CYAN}└─ Sem Procedimento Complementar{Cores.RESET}")
        print(f"     {Cores.CYAN}└─ Fluxo direto para anexos{Cores.RESET}\n")
        
        print(f"  {Cores.VERMELHO}0.{Cores.RESET} {Cores.BOLD}← Voltar ao menu principal{Cores.RESET}")
        
        print(f"\n{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
        
        escolha = input(f"{Cores.BOLD}Digite sua escolha: {Cores.RESET}").strip()
        
        if escolha == "1":
            return "completo"
        elif escolha == "2":
            return "simplificado"
        elif escolha == "0":
            return None
        else:
            exibir_mensagem_erro("Opção inválida! Digite 1, 2 ou 0.")
            time.sleep(1.5)

def exibir_informacoes_sistema(pasta_base):
    """Exibe informações sobre o sistema."""
    exibir_banner()
    print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}")
    print(f"{Cores.BOLD}{'ℹ️  INFORMAÇÕES DO SISTEMA'.center(80)}{Cores.RESET}")
    print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
    
    print(f"{Cores.BOLD}📌 Versão:{Cores.RESET} 2.0")
    print(f"{Cores.BOLD}🏥 Sistema:{Cores.RESET} Automação Porto Seguro Odonto")
    print(f"{Cores.BOLD}📅 Ano:{Cores.RESET} {datetime.now().year}")
    print(f"{Cores.BOLD}📁 Diretório:{Cores.RESET} {pasta_base}")
    
    pasta_logs = pasta_base / "logs"
    logs_falha, logs_sucesso = listar_logs_disponiveis(pasta_logs)
    print(f"\n{Cores.BOLD}📊 Estatísticas:{Cores.RESET}")
    print(f"  • Total de logs de falha: {len(logs_falha)}")
    print(f"  • Total de logs de sucesso: {len(logs_sucesso)}")
    
    pasta_arquivos = pasta_base / "arquivos"
    if pasta_arquivos.exists():
        total_arquivos = len(list(pasta_arquivos.rglob('*.*')))
        print(f"  • Arquivos na pasta 'arquivos': {total_arquivos}")
    
    print(f"\n{Cores.BOLD}📋 Procedimentos Cadastrados:{Cores.RESET} {len(MAPEAMENTO_PROCEDIMENTOS)}")
    
    print(f"\n{Cores.BOLD}👥 Tipos de Usuário:{Cores.RESET}")
    print(f"  • Usuário Completo (00281027) - Com procedimento complementar")
    print(f"  • Usuário Simplificado (00260115) - Sem procedimento complementar")
    
    print(f"\n{Cores.BOLD}🔧 Funcionalidades:{Cores.RESET}")
    print(f"  ✓ Processamento inteligente de guias")
    print(f"  ✓ Anexo automático de arquivos")
    print(f"  ✓ Preenchimento de observações")
    print(f"  ✓ Sistema de logs detalhado")
    print(f"  ✓ Delays estratégicos anti-detecção")
    print(f"  ✓ Seleção de tipo de usuário")
    
    print(f"\n{Cores.CYAN}{'═' * 80}{Cores.RESET}")
    input(f"\n{Cores.CYAN}Pressione ENTER para voltar...{Cores.RESET}")

# ======================== PONTO DE ENTRADA DO SCRIPT ========================
if __name__ == "__main__":
    pasta_base = Path(__file__).parent
    pasta_arquivos = pasta_base / "arquivos"
    pasta_arquivos.mkdir(exist_ok=True)
    
    # Loop do menu principal
    while True:
        try:
            exibir_menu_principal()
            escolha = input(f"{Cores.BOLD}Digite sua escolha: {Cores.RESET}").strip()
            
            if escolha == "1":
                # Iniciar processamento
                # Primeiro, seleciona o tipo de usuário
                tipo_usuario = selecionar_tipo_usuario()
                if tipo_usuario is None:
                    continue  # Usuário cancelou, volta ao menu principal
                
                lista_planilhas = glob.glob(str(pasta_base / "*.xlsx"))
                if not lista_planilhas:
                    exibir_banner()
                    exibir_mensagem_erro("ERRO: Nenhuma planilha (.xlsx) foi encontrada na pasta do robô.")
                    exibir_mensagem_info(f"Pasta esperada: {pasta_base}")
                    input("\nPressione ENTER para voltar ao menu...")
                    continue
                elif len(lista_planilhas) > 1:
                    exibir_banner()
                    exibir_mensagem_erro("ERRO: Mais de uma planilha (.xlsx) foi encontrada. Deixe apenas uma.")
                    exibir_mensagem_info("Planilhas encontradas:")
                    for pl in lista_planilhas:
                        print(f"  • {Path(pl).name}")
                    input("\nPressione ENTER para voltar ao menu...")
                    continue
                
                caminho_planilha = lista_planilhas[0]
                exibir_banner()
                exibir_mensagem_sucesso(f"Planilha encontrada: {Path(caminho_planilha).name}")
                
                # Exibe o usuário selecionado
                bot_temp = PortoSeguroBot(tipo_usuario)
                print(f"\n{Cores.CYAN}👤 Usuário selecionado:{Cores.RESET} {Cores.BOLD}{bot_temp.usuarios[tipo_usuario]['nome']}{Cores.RESET}")
                print(f"{Cores.CYAN}🔑 Login:{Cores.RESET} {bot_temp.login}")
                
                # Confirmação antes de iniciar
                print(f"\n{Cores.AMARELO}⚠ ATENÇÃO:{Cores.RESET} O processamento automático será iniciado.")
                confirmacao = input(f"{Cores.BOLD}Deseja continuar? (S/N): {Cores.RESET}").strip().upper()
                
                if confirmacao != 'S':
                    exibir_mensagem_info("Processamento cancelado pelo usuário.")
                    time.sleep(1.5)
                    continue
                
                bot = PortoSeguroBot(tipo_usuario)
                bot.log_manager = LogManager(pasta_base)
                
                try:
                    processar_guias_da_planilha(bot, caminho_planilha, str(pasta_arquivos))
                except Exception as e:
                    exibir_mensagem_erro(f"Ocorreu um erro fatal na execução: {e}")
                    bot.log_manager.registrar_falha("FATAL", "Erro crítico não tratado", traceback.format_exc())
                finally:
                    bot.log_manager.gerar_resumos()
                    if bot.driver:
                        exibir_mensagem_info("Fechando navegador...")
                        bot.driver.quit()
                    
                    print(f"\n{Cores.VERDE}{'─' * 80}{Cores.RESET}")
                    input(f"{Cores.BOLD}Pressione ENTER para voltar ao menu principal...{Cores.RESET}")
            
            elif escolha == "2":
                # Visualizar logs
                menu_visualizar_logs(pasta_base)
            
            elif escolha == "3":
                # Informações do sistema
                exibir_informacoes_sistema(pasta_base)
            
            elif escolha == "0":
                # Sair
                exibir_banner()
                print(f"\n{Cores.CYAN}{'═' * 80}{Cores.RESET}")
                print(f"{Cores.BOLD}{'👋 Obrigado por usar o Sistema de Automação Porto Seguro!'.center(80)}{Cores.RESET}")
                print(f"{Cores.CYAN}{'═' * 80}{Cores.RESET}\n")
                break
            
            else:
                exibir_mensagem_erro("Opção inválida! Digite um número válido.")
                time.sleep(1.5)
        
        except KeyboardInterrupt:
            exibir_banner()
            print(f"\n{Cores.AMARELO}⚠ Operação interrompida pelo usuário.{Cores.RESET}")
            confirmacao = input(f"{Cores.BOLD}Deseja realmente sair? (S/N): {Cores.RESET}").strip().upper()
            if confirmacao == 'S':
                print(f"\n{Cores.CYAN}Até logo! 👋{Cores.RESET}\n")
                break
        except Exception as e:
            exibir_mensagem_erro(f"Erro inesperado no menu: {e}")
            time.sleep(2)