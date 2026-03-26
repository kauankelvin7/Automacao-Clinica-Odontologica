import os
import sys
import traceback
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

from bot import LogManager, PortoSeguroBot
from constants import MAPEAMENTO_PROCEDIMENTOS, Cores
from utils import (exibir_banner, exibir_mensagem_alerta, exibir_mensagem_erro,
                   exibir_mensagem_info, exibir_mensagem_sucesso)

# Carrega variáveis de ambiente do arquivo .env (um nível acima da pasta src/)
BASE_DIR = Path(__file__).parent.parent
load_dotenv(BASE_DIR / ".env")

# ======================== FUNÇÃO DE ORQUESTRAÇÃO PRINCIPAL ========================
def processar_guias_da_planilha(bot, caminho_planilha, pasta_arquivos_busca):
    """Orquestra todo o processo de automação, desde a leitura da planilha até o final."""
    try:
        workbook = openpyxl.load_workbook(caminho_planilha)
        sheet = workbook.active
    except FileNotFoundError:
        exibir_mensagem_erro(f"Planilha não encontrada: {caminho_planilha}")
        return

    guias_a_processar = []
    print(f"\n{Cores.CYAN}Lendo planilha... Total de linhas: {sheet.max_row}{Cores.RESET}")
    
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_col=4, values_only=True), 1):
        if row[0]:
            cod_simplificado = ""
            if row[1]:
                cod_raw = str(row[1]).strip()
                if '.' in cod_raw and cod_raw.replace('.', '').replace('-', '').isdigit():
                    cod_simplificado = str(int(float(cod_raw)))
                else:
                    cod_simplificado = cod_raw
            
            if cod_simplificado and cod_simplificado not in MAPEAMENTO_PROCEDIMENTOS:
                exibir_mensagem_erro(f"ERRO na linha {i} da planilha: Código '{cod_simplificado}' não foi encontrado no dicionário.")
                continue

            guia_raw = str(row[0]).strip()
            if '.' in guia_raw and guia_raw.replace('.', '').replace('-', '').isdigit():
                guia_formatada = str(int(float(guia_raw)))
            else:
                guia_formatada = guia_raw
            
            dente_raw = str(row[2]).strip() if row[2] else ""
            if dente_raw and '.' in dente_raw and dente_raw.replace('.', '').replace('-', '').isdigit():
                dente_formatado = str(int(float(dente_raw)))
            else:
                dente_formatado = dente_raw

            guia_dict = {
                'guia': guia_formatada,
                'cod_simplificado': cod_simplificado,
                'dente': dente_formatado,
                'observacao': str(row[3]).strip() if row[3] else ""
            }
            guias_a_processar.append(guia_dict)
    
    print(f"{Cores.VERDE}Total de guias válidas encontradas: {len(guias_a_processar)}{Cores.RESET}")
    
    total_guias = len(guias_a_processar)
    exibir_banner()
    print(f"\n{Cores.BOLD}{Cores.AZUL}{'═' * 80}{Cores.RESET}")
    print(f"{Cores.BOLD}INICIANDO PROCESSAMENTO DE {total_guias} GUIAS{Cores.RESET}")
    
    sucessos = 0
    try:
        bot.iniciar_driver()
        bot.fazer_login()
    except Exception as e:
        exibir_mensagem_erro(f"Falha crítica na inicialização: {e}")
        bot.log_manager.registrar_falha("GERAL", "Falha no Login/Inicialização", traceback.format_exc())
        return

    for i, guia_info in enumerate(guias_a_processar, 1):
        id_guia = guia_info['guia']
        cod_simplificado = guia_info['cod_simplificado']
        dente_val = guia_info['dente']

        proc_info = MAPEAMENTO_PROCEDIMENTOS.get(cod_simplificado, {})
        proc_code = proc_info.get('codigo_completo', "")
        proc_desc = proc_info.get('descricao', "")
        
        print(f"\n{Cores.MAGENTA}{'─' * 80}")
        print(f"► Processando Guia {i}/{total_guias}: {Cores.BOLD}{id_guia}{Cores.RESET}")
        
        try:
            bot.navegar_meus_tratamentos()
            if not bot.buscar_guia(id_guia):
                bot.log_manager.registrar_falha(id_guia, "Guia não encontrada")
                continue
            
            if bot.verificar_guia_ja_finalizada():
                exibir_mensagem_sucesso("Guia já completamente finalizada. Pulando.")
                sucessos += 1
                bot.log_manager.registrar_sucesso(id_guia, "N/A", "N/A - Já finalizada")
                continue

            procedimento_ja_confirmado = bot.verificar_procedimento_ja_confirmado()
            
            if not procedimento_ja_confirmado:
                observacao = guia_info.get('observacao', "")
                if not bot.confirmar_realizacao_procedimento(proc_code, proc_desc, dente_val, observacao, id_guia):
                    bot.log_manager.registrar_falha(id_guia, "Falha ao confirmar procedimento ou na validação do modal")
                    continue
            else:
                exibir_mensagem_sucesso("Procedimento já confirmado ($ laranja). Indo direto para anexos...")
                bot._pause_between_actions(3)

            lista_de_anexos = bot.buscar_arquivos_guia(id_guia, pasta_arquivos_busca)
            if not bot.anexar_todos_os_arquivos(lista_de_anexos):
                bot.log_manager.registrar_falha(id_guia, "Falha no anexo de arquivos")
                continue

            if not bot.submeter_para_pagamento():
                bot.log_manager.registrar_falha(id_guia, "Falha na submissão final")
                continue
            
            print(f"\n{Cores.BG_VERDE} ✓ GUIA {id_guia} FINALIZADA COM SUCESSO {Cores.RESET}\n")
            sucessos += 1
            bot.log_manager.registrar_sucesso(id_guia, cod_simplificado, dente_val)

        except KeyboardInterrupt:
            exibir_mensagem_alerta("\nProcessamento interrompido pelo usuário!")
            break
        except Exception as e:
            exibir_mensagem_erro(f"Erro inesperado na guia {id_guia}: {e}")
            bot.log_manager.registrar_falha(id_guia, "Erro inesperado", traceback.format_exc())

    print(f"\n{Cores.CYAN}{'═' * 80}\n{'PROCESSAMENTO FINALIZADO'.center(80)}\n{'═' * 80}{Cores.RESET}")
    print(f"{Cores.BOLD}Total de guias na planilha:{Cores.RESET} {total_guias}")
    print(f"{Cores.VERDE}Processadas com sucesso:{Cores.RESET} {sucessos}")
    print(f"{Cores.VERMELHO}Guias com falha:{Cores.RESET} {len(bot.log_manager.falhas)}")

# ======================== FUNÇÃO PRINCIPAL E MENU ========================
def main():
    pasta_base = BASE_DIR
    pasta_arquivos = pasta_base / "arquivos"
    
    exibir_banner()
    
    while True:
        print(f"{Cores.BOLD}--- MENU PRINCIPAL ---{Cores.RESET}")
        print("[1] Iniciar Processamento (Usuário Completo)")
        print("[2] Iniciar Processamento (Usuário Simplificado)")
        print("[3] Visualizar Logs")
        print("[0] Sair")
        
        opcao = input(f"\n{Cores.AMARELO}Escolha uma opção: {Cores.RESET}")
        
        if opcao == "1":
            bot = PortoSeguroBot(tipo_usuario="completo")
            bot.log_manager = LogManager(pasta_base)
            planilha = next(pasta_base.glob("*.xlsx"), None)
            if planilha:
                processar_guias_da_planilha(bot, str(planilha), str(pasta_arquivos))
                bot.log_manager.gerar_resumos()
                if bot.driver: bot.driver.quit()
            else:
                exibir_mensagem_erro("Nenhuma planilha .xlsx encontrada na pasta raiz.")
        
        elif opcao == "2":
            bot = PortoSeguroBot(tipo_usuario="simplificado")
            bot.log_manager = LogManager(pasta_base)
            planilha = next(pasta_base.glob("*.xlsx"), None)
            if planilha:
                processar_guias_da_planilha(bot, str(planilha), str(pasta_arquivos))
                bot.log_manager.gerar_resumos()
                if bot.driver: bot.driver.quit()
            else:
                exibir_mensagem_erro("Nenhuma planilha .xlsx encontrada na pasta raiz.")
        
        elif opcao == "3":
            pasta_logs = pasta_base / "logs"
            logs = sorted(pasta_logs.glob("*.txt"), reverse=True)
            if logs:
                print(f"\n{Cores.CYAN}--- LOGS RECENTES ---{Cores.RESET}")
                for i, log in enumerate(logs[:5], 1):
                    print(f"[{i}] {log.name}")
                print("[0] Voltar")
                op_log = input(f"\n{Cores.AMARELO}Escolha um log para ver: {Cores.RESET}")
                if op_log.isdigit() and 0 < int(op_log) <= len(logs):
                    os.system(f'notepad "{logs[int(op_log)-1]}"' if os.name == 'nt' else f'cat "{logs[int(op_log)-1]}"')
            else:
                exibir_mensagem_info("Nenhum log encontrado.")
        
        elif opcao == "0":
            print(f"\n{Cores.VERDE}Até logo!{Cores.RESET}")
            break
        else:
            exibir_mensagem_alerta("Opção inválida.")

if __name__ == "__main__":
    main()